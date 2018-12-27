from PyQt5.QtCore import QTimer
import sqlite3
import time
from openpyxl import load_workbook
import os
import re
class SqlDataHandle(object):
    total = 0
    def __init__(self,name="data.db",path="backlog.xlsx"):
        self.dbname=name
        self.pathexcl=path
        self.tablesql="create table test(a varchar(0,9999999999),b varchar(0,9999999999),c varchar(0,9999999999),d varchar(0,9999999999),e varchar(0,9999999999))"
        self.formatsql='insert into test values(?,?,?,?,?)'
        self.header =None
        self.tablepage=0
        self.Parseiterationlist=[]
    def eachXlsx(self,pathexcl=None,MangyParserHandler=False,name=None):
        __header=[]
        if pathexcl is None:
            __datasql = self.pathexcl
        else:
            if str(pathexcl).endswith(".xlsx") is not False:
                __datasql=self.pathexcl
        if MangyParserHandler is False:
            wb = load_workbook(__datasql)
            ws = wb.worksheets[0]
            for index, row in enumerate(ws.rows):
                if index == 0:
                    self.header=tuple(map(lambda x: x.value, row))
                    continue
                yield tuple(map(lambda x: x.value, row))
        else:
            wb = load_workbook(__datasql)
            ws = wb.get_sheet_by_name(name)
            for index, row in enumerate(ws.rows):
                if index == 0:
                    self.header = tuple(map(lambda x: x.value, row))
                    continue
                yield tuple(map(lambda x: x.value, row))
    def insertdatasql (self,sqltable=None,tablepage=0):
        if tablepage==0:
            if sqltable is None:
                sqltable = self.tablesql
            conn = sqlite3.connect(self.dbname)
            cursor = conn.cursor()
            try:
                cursor.execute(sqltable)
            except sqlite3.OperationalError:
                print("table test already exists")
            finally:
                print(self.formatsql)
                cursor.executemany(self.formatsql, self.eachXlsx())
                conn.commit()
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
        else:
            if sqltable is None:
                sqltable = self.tablesql
            if tablepage >= self.tablepage:
                tablepage=self.tablepage
            conn = sqlite3.connect(self.dbname)
            cursor = conn.cursor()
            for x in range(tablepage):
                try:
                    cursor.execute(sqltable[x])
                except sqlite3.OperationalError:
                    print("table test already exists")
                finally:
                    print(self.formatsql[x])
                    cursor.executemany(self.formatsql[x], self.Parseiterationlist[x])
                    conn.commit()
            if cursor:
                cursor.close()
            if conn:
                conn.close()


    def parsedata(self,*agre):
        self.eachXlsx().__next__()
        __Tsql=[]
        __Dsql = []
        z, f = os.path.split(self.pathexcl)
        tablename =f[0:-5]

        for x in self.header:
            __Tsql.append(("%s varchar(0, 80000)" % x))
            __Dsql.append("?")
        self.tablesql= str(("create table %s (%s)" % (tablename, (str(__Tsql).replace("'", ""))))).replace("[", "").replace("]", "")
        self.formatsql=str("insert into %s values(%s)"%(tablename, str(__Dsql).replace("'", ""))).replace("[", "").replace("]", "")
        return self.tablesql
    def manyparsedata(self,onlygetshettlnames=False):
        wb = load_workbook(self.pathexcl)
        names = wb.get_sheet_names()
        if onlygetshettlnames is True:
            return  names
        self.tablepage=len(names)
        self.tablesql=[]
        self.formatsql=[]
        for tablename in names:
            ws=wb.get_sheet_by_name(tablename)
            for index, row in enumerate(ws.rows):
                if index == 0:
                    self.header = tuple(map(lambda x: x.value, row))
                    __Tsql=[]
                    __Dsql = []
                    for x in self.header:
                        __Tsql.append(("%s varchar(0, 80000)" % x))
                        __Dsql.append("?")
                    tablesql= str(("create table %s (%s)" % (tablename, (str(__Tsql).replace("'", ""))))).replace("[", "").replace("]", "")
                    self.tablesql.append(tablesql)
                    formatsql=str("insert into %s values(%s)"%(tablename, str(__Dsql).replace("'", ""))).replace("[", "").replace("]", "")
                    self.formatsql.append(formatsql)
    def manyinsertdatasql(self,pagecount=30):
        self.__manyeachXlsx()
        self.manyparsedata()
        self.insertdatasql(tablepage=int(pagecount))
    def __manyeachXlsx(self):
        __Parseiterationlist=[]
        for x in self.manyparsedata(onlygetshettlnames=True):
            __Parseiterationlist.append(self.eachXlsx(MangyParserHandler=True, name=x))
        self.Parseiterationlist = __Parseiterationlist
        return self.Parseiterationlist
    def timer(self,fun):
        start = time()
        fun()
        delta = time() - start
        print('导入用时：', delta)
    def __call__(self, *args, **kwargs):
        print("将对象变成可回调")
        return self
    def default_run(self):
        self.insertdatasql(self.parsedata())




if __name__=="__main__":
    testsql=SqlDataHandle()
    testsql.manyinsertdatasql()