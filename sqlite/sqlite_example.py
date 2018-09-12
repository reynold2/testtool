from random import choice, randrange
from string import digits, ascii_letters
from os import listdir
import sqlite3
from time import time
from openpyxl import Workbook, load_workbook


total = 0
def generateRandomData():
    #total表示记录总条数
    global total
    characters = digits+ascii_letters
    for i in range(50):
        xlsName = str(i)+'.xlsx'
        #随机数，每个xlsx文件的行数不一样
        totalLines = randrange(10**5)
        wb = Workbook()
        ws = wb.worksheets[0]
        #表头
        ws.append(['a', 'b', 'c', 'd', 'e'])
        #随机数据，每行5个字段，每个字段30个字符
        for j in range(totalLines):
            line = [''.join((choice(characters) for ii in range(30))) for jj in range(5)]
            ws.append(line)
            total += 1
        #保存xlsx文件
        wb.save(xlsName)



#针对每个xlsx文件的生成器
def eachXlsx(xlsxFn):
    wb = load_workbook(xlsxFn)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows):
        #忽略表头
        if index == 0:
            continue
        yield tuple(map(lambda x:x.value, row))



#导入
def xlsx2sqlite():
    #获取所有xlsx文件
    xlsxs=('0.xlsx','1.xlsx','2.xlsx')
    # xlsxs = (fn for fn in listdir('xlsxs'))
    #连接数据库，创建游标
    conn = sqlite3.connect('data2.db')

    cur = conn.cursor()
    cur.execute("create table test(a varchar(0,9999999999),b varchar(0,9999999999),c varchar(0,9999999999),d varchar(0,9999999999),e varchar(0,9999999999))")
    for xlsx in xlsxs:
        #批量导入，减少提交事务的次数，可以提高速度
        sql = 'insert into test values(?,?,?,?,?)'
        cur.executemany(sql, eachXlsx(xlsx))
        conn.commit()







# generateRandomData()

start = time()

delta = time()-start
print('导入用时：', delta)
xlsx2sqlite()
print('导入速度（条/秒）：', total/delta)